import os
from datetime import datetime
from io import BytesIO, StringIO

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.core.files.storage import FileSystemStorage
from django.core.management import call_command
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.dashboard.services.alertas_service import (
    LIMITE_SUGERENCIAS_ALERTAS,
    MIN_CARACTERES_BUSQUEDA_ALERTAS,
    buscar_amids_alertas,
    buscar_ubicaciones_alertas,
    obtener_contexto_alertas,
    obtener_ubicaciones_alertas_disponibles,
    obtener_alertas_para_exportar,
)
from apps.dashboard.services.oracle_connection import obtener_conexion_oracle
from apps.dashboard.services.preferencias_alertas_service import (
    guardar_preferencias_alertas_usuario,
    obtener_preferencias_alertas_usuario,
)
from apps.dashboard.services.reglas_alertas_service import (
    actualizar_reglas_alertas,
    iniciar_recalculo_en_segundo_plano,
    leer_log_recalculo,
    obtener_editor_reglas_alertas,
    recalculo_en_curso,
    usuario_puede_editar_reglas,
)

from .models import LogImportacion
from .services.baterias_service import (
    construir_tabla_bateria,
    obtener_ahora_referencia,
    obtener_bloques_bateria_oracle,
    obtener_contexto_baterias,
    obtener_detalle_caidas_bateria_oracle,
    obtener_rango_fechas_panel,
)
from .services.gps_service import obtener_contexto_gps

def usuario_es_admin(user):
    return user.is_superuser or user.groups.filter(name="Admin").exists()


@login_required
def ejecutar_comando_admin(request):
    if not usuario_es_admin(request.user):
        messages.error(request, "No tienes permisos para ejecutar acciones administrativas.")
        return redirect("dashboard:panel_perfil")

    if request.method != "POST":
        return redirect("dashboard:panel_perfil")

    accion = request.POST.get("accion")
    salida = StringIO()

    acciones_antiguas_sqlite = [
        "actualizar_validadores",
        "importar_oracle_2h",
        "importar_oracle_14d",
        "cargar_limpios",
        "limpiar_antiguos",
    ]

    try:
        if accion == "probar_oracle":
            call_command("probar_oracle", stdout=salida, stderr=salida)

        elif accion == "importar_ubicaciones":
            archivo = request.FILES.get("archivo_version_zp")

            if not archivo:
                messages.error(request, "Debes seleccionar un archivo Excel.")
                return redirect("dashboard:panel_perfil")

            carpeta_tmp = os.path.join(settings.BASE_DIR, "temp_uploads")
            os.makedirs(carpeta_tmp, exist_ok=True)

            storage = FileSystemStorage(location=carpeta_tmp)
            nombre_archivo = storage.save(archivo.name, archivo)
            ruta_archivo = storage.path(nombre_archivo)

            try:
                call_command(
                    "importar_ubicaciones_esperadas",
                    ruta_archivo,
                    stdout=salida,
                    stderr=salida,
                )
            finally:
                try:
                    os.remove(ruta_archivo)
                except OSError:
                    pass

        elif accion in acciones_antiguas_sqlite:
            messages.warning(
                request,
                "Esta acción pertenece al flujo antiguo SQLite y está deshabilitada. "
                "Los datos operativos ahora se consultan directamente desde Oracle."
            )
            return redirect("dashboard:panel_perfil")

        else:
            messages.error(request, "Acción no reconocida.")
            return redirect("dashboard:panel_perfil")

        request.session["resultado_comando_admin"] = salida.getvalue()
        messages.success(request, "Proceso ejecutado correctamente.")

    except Exception as error:
        request.session["resultado_comando_admin"] = salida.getvalue()
        messages.error(request, f"Error ejecutando proceso: {error}")

    return redirect("dashboard:panel_perfil")


@login_required
def buscar_exclusiones_alertas(request):
    """Endpoint liviano para los autocompletados del panel de alertas."""

    if request.method != "GET":
        return JsonResponse({"detalle": "Metodo no permitido."}, status=405)

    tipo = request.GET.get("tipo", "").strip().lower()
    termino = request.GET.get("q", "").strip()[:120]

    if len(termino) < MIN_CARACTERES_BUSQUEDA_ALERTAS:
        return JsonResponse(
            {
                "resultados": [],
                "minimo_caracteres": MIN_CARACTERES_BUSQUEDA_ALERTAS,
            }
        )

    if tipo == "amid":
        valores = buscar_amids_alertas(termino, LIMITE_SUGERENCIAS_ALERTAS)
    elif tipo == "ubicacion":
        valores = buscar_ubicaciones_alertas(termino, LIMITE_SUGERENCIAS_ALERTAS)
    else:
        return JsonResponse({"detalle": "Tipo de busqueda no valido."}, status=400)

    return JsonResponse(
        {
            "resultados": [
                {
                    "valor": valor,
                    "etiqueta": valor,
                }
                for valor in valores
            ],
            "minimo_caracteres": MIN_CARACTERES_BUSQUEDA_ALERTAS,
            "limite": LIMITE_SUGERENCIAS_ALERTAS,
        }
    )


@login_required
def detalle_caidas_bateria(request):
    """Devuelve las caídas de un AMID solo cuando el usuario abre el detalle."""
    if request.method != "GET":
        return JsonResponse({"detalle": "Metodo no permitido."}, status=405)

    amid = request.GET.get("amid", "").strip()
    if not amid or not amid.isdigit():
        return JsonResponse(
            {"detalle": "Debes indicar un AMID valido."},
            status=400,
        )

    try:
        resultado = obtener_detalle_caidas_bateria_oracle(amid=amid, dias=14)
    except Exception:
        return JsonResponse(
            {"detalle": "El detalle de caidas de Oracle no esta disponible."},
            status=503,
        )

    return JsonResponse(resultado)


@login_required
def panel_alertas(request):
    if request.method == "POST":
        # Se obtiene el catalogo completo solo al guardar para conservar la
        # validacion existente. En los GET ya no se envia al HTML.
        ubicaciones_disponibles = obtener_ubicaciones_alertas_disponibles()
        limpiar_preferencias = request.POST.get("limpiar_preferencias") == "1"
        texto_amids = (
            ""
            if limpiar_preferencias
            else request.POST.get("amids_excluidos", "")
        )
        ubicaciones = (
            []
            if limpiar_preferencias
            else request.POST.getlist("ubicaciones_excluidas")
        )

        try:
            preferencias = guardar_preferencias_alertas_usuario(
                usuario=request.user,
                texto_amids=texto_amids,
                ubicaciones=ubicaciones,
                ubicaciones_disponibles=ubicaciones_disponibles,
            )
            if limpiar_preferencias:
                messages.success(request, "Preferencias de alertas restablecidas.")
            else:
                messages.success(
                    request,
                    "Preferencias guardadas: "
                    f"{len(preferencias['amids_excluidos'])} AMID y "
                    f"{len(preferencias['ubicaciones_excluidas'])} ubicaciones excluidas.",
                )
        except ValueError as error:
            messages.error(request, str(error))

        return redirect("dashboard:panel_alertas")

    preferencias = obtener_preferencias_alertas_usuario(request.user)

    try:
        contexto = obtener_contexto_alertas(request, preferencias=preferencias)
    except ValueError as error:
        messages.error(request, str(error))
        return redirect("dashboard:panel_alertas")

    contexto.update(
        {
            "preferencias_alertas": preferencias,
            "amids_excluidos_texto": ",".join(
                str(amid) for amid in preferencias["amids_excluidos"]
            ),
            "total_amids_excluidos": len(preferencias["amids_excluidos"]),
            "total_ubicaciones_excluidas": len(
                preferencias["ubicaciones_excluidas"]
            ),
            "total_preferencias_alertas": (
                len(preferencias["amids_excluidos"])
                + len(preferencias["ubicaciones_excluidas"])
            ),
        }
    )
    return render(request, "dashboard/panel_alertas.html", contexto)


COLUMNAS_EXCEL_ALERTAS = (
    ("Prioridad", "nivel_alerta_global"),
    ("AMID", "amid"),
    ("Ubicación actual", "ubicacion_actual"),
    ("Último estatus", "ultimo_estatus"),
    ("Estado del estatus", "texto_estatus"),
    ("GPS", "nivel_alerta_gps"),
    ("Motivo GPS", "motivo_alerta_gps"),
    ("Registros GPS hoy", "gps_total_hoy"),
    ("GPS 0,0 hoy", "gps_cero_hoy"),
    ("GPS 0,0 histórico", "gps_cero_hist"),
    ("GPS 0,0 % hoy", "gps_cero_porc_hoy"),
    ("GPS 0,0 % histórico", "gps_cero_porc_hist"),
    ("Racha máxima GPS 0,0", "racha_max_gps_cero"),
    ("Batería", "nivel_alerta_bateria"),
    ("Motivo batería", "motivo_alerta_bateria"),
    ("Batería actual (%)", "bateria_actual"),
    ("Caídas hoy", "caidas_hoy"),
    ("Caídas históricas", "caidas_hist"),
    ("Caída máxima hoy", "caida_max_hoy"),
    ("Caída máxima histórica", "caida_max_hist"),
    ("Batería 0 hoy", "bateria_cero_hoy"),
    ("Batería 0 histórica", "bateria_cero_hist"),
    ("Motivo principal", "motivo_principal"),
    ("Acción sugerida", "accion_sugerida"),
    ("Fecha actualización", "fecha_actualizacion"),
)


def crear_excel_alertas(alertas, fecha_generacion=None):
    """Construye la hoja formateada con todos los AMID activos."""

    fecha_generacion = fecha_generacion or obtener_ahora_referencia()
    wb = Workbook()
    wb.properties.title = "Panel de Alertas"
    wb.properties.subject = "Alertas vigentes de validadores activos"

    ws = wb.active
    ws.title = "Panel de Alertas"
    ultima_columna = get_column_letter(len(COLUMNAS_EXCEL_ALERTAS))

    ws.merge_cells(f"A1:{ultima_columna}1")
    ws["A1"] = "Panel de Alertas - Validadores activos"
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{ultima_columna}2")
    ws["A2"] = (
        f"Exportado: {fecha_generacion.strftime('%d-%m-%Y %H:%M:%S')} | "
        f"Total AMID activos: {len(alertas)} | Sin filtros ni exclusiones de usuario"
    )
    ws["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A2"].font = Font(color="1F2937", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    fila_encabezados = 4
    ws.append([])
    ws.append([titulo for titulo, _ in COLUMNAS_EXCEL_ALERTAS])

    for alerta in alertas:
        fila = []
        for _, clave in COLUMNAS_EXCEL_ALERTAS:
            valor = alerta.get(clave)
            if clave == "amid" and valor is not None:
                valor = str(valor)
            fila.append(preparar_valor_excel(valor))
        ws.append(fila)

    fill_encabezado = PatternFill("solid", fgColor="1F4E78")
    borde_encabezado = Border(
        bottom=Side(style="thin", color="B4C7E7")
    )
    for celda in ws[fila_encabezados]:
        celda.fill = fill_encabezado
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        celda.border = borde_encabezado
    ws.row_dimensions[fila_encabezados].height = 34

    if alertas:
        tabla = Table(
            displayName="TablaPanelAlertas",
            ref=f"A{fila_encabezados}:{ultima_columna}{ws.max_row}",
        )
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabla)

    colores_nivel = {
        "CRITICA": ("F8D7DA", "842029"),
        "ALTA": ("FCE5CD", "9C5700"),
        "ADVERTENCIA": ("FFF3CD", "664D03"),
        "OK": ("D1E7DD", "0F5132"),
    }
    columnas_nivel = (1, 6, 14)
    colores_estatus = {
        "Con estatus": ("D1E7DD", "0F5132"),
        "Hace más de 1 hora": ("FFF3CD", "664D03"),
        "Sin estatus hoy": ("F8D7DA", "842029"),
    }

    for numero_fila in range(fila_encabezados + 1, ws.max_row + 1):
        for numero_columna in columnas_nivel:
            celda = ws.cell(numero_fila, numero_columna)
            colores = colores_nivel.get(str(celda.value or "").upper())
            if colores:
                celda.fill = PatternFill("solid", fgColor=colores[0])
                celda.font = Font(color=colores[1], bold=True)
            celda.alignment = Alignment(horizontal="center", vertical="center")

        celda_estatus = ws.cell(numero_fila, 5)
        colores = colores_estatus.get(str(celda_estatus.value or ""))
        if colores:
            celda_estatus.fill = PatternFill("solid", fgColor=colores[0])
            celda_estatus.font = Font(color=colores[1], bold=True)

        ws.cell(numero_fila, 2).number_format = "@"
        for numero_columna in (4, 25):
            ws.cell(numero_fila, numero_columna).number_format = "dd-mm-yyyy hh:mm:ss"
        for numero_columna in (11, 12, 16, 19, 20):
            ws.cell(numero_fila, numero_columna).number_format = "0.00"

    anchos = {
        "A": 14, "B": 13, "C": 38, "D": 20, "E": 20,
        "F": 14, "G": 38, "H": 18, "I": 14, "J": 18,
        "K": 16, "L": 20, "M": 22, "N": 14, "O": 40,
        "P": 18, "Q": 13, "R": 18, "S": 18, "T": 22,
        "U": 16, "V": 20, "W": 40, "X": 28, "Y": 20,
    }
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    for fila in ws.iter_rows(
        min_row=fila_encabezados + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(COLUMNAS_EXCEL_ALERTAS),
    ):
        for celda in fila:
            celda.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 85
    ws.auto_filter.ref = f"A{fila_encabezados}:{ultima_columna}{ws.max_row}"
    ws.print_title_rows = f"1:{fila_encabezados}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return wb


@login_required
def exportar_alertas_excel(request):
    """Exporta todos los AMID activos, sin filtros ni preferencias de usuario."""

    alertas = obtener_alertas_para_exportar()
    fecha_generacion = obtener_ahora_referencia()
    wb = crear_excel_alertas(alertas, fecha_generacion=fecha_generacion)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nombre_archivo = (
        "panel_alertas_"
        f"{fecha_generacion.strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{nombre_archivo}"'
    )
    return response

@login_required
def panel_baterias(request):
    contexto = obtener_contexto_baterias(request)
    contexto["active_page"] = "baterias"
    return render(request, "dashboard/panel_baterias.html", contexto)


@login_required
def panel_gps(request):
    contexto = obtener_contexto_gps(request)
    contexto["active_page"] = "gps"
    return render(request, "dashboard/panel_gps.html", contexto)


@login_required
def exportar_gps_excel(request):
    """
    Exporta el Excel estándar del AMID desde el botón GPS.

    Por ahora, igual que Baterías:
    - siempre exporta 14 días completos
    - no usa filtros de fecha/hora del panel GPS
    """

    amid = request.GET.get("amid", "").strip()

    if not amid:
        return HttpResponse("Debe indicar un AMID para exportar.", status=400)

    dias = 14
    hora_inicio = "00:00"
    hora_fin = "23:30"

    try:
        wb = crear_excel_completo_amid(
            amid=amid,
            dias=dias,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
        )
    except ValueError:
        return HttpResponse("El AMID ingresado no es válido.", status=400)
    except Exception as error:
        return HttpResponse(
            f"Error consultando datos en Oracle: {error}",
            status=500,
        )

    if wb is None:
        return HttpResponse(
            f"No existen registros para el AMID {amid} en los últimos 14 días.",
            status=404,
        )

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    fecha_exportacion = obtener_ahora_referencia().strftime("%Y%m%d_%H%M%S")
    filename = f"datos_amid_{amid}_14_dias_{fecha_exportacion}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def obtener_registros_completos_oracle(amid, fecha_inicio, fecha_fin):
    """
    Obtiene todos los datos relevantes del AMID desde Oracle
    para la hoja 'Registros completos'.
    """

    fecha_inicio_texto = fecha_inicio.strftime("%Y-%m-%d %H:%M:%S")
    fecha_fin_texto = fecha_fin.strftime("%Y-%m-%d %H:%M:%S")

    query = """
        SELECT
            ID,
            AMID,
            FEC_DESCARGA,
            FEC_ESTADO,
            BUSID,
            OP,
            VERSION,
            PATENTE,
            TD01,
            TD04,
            TABLA,
            VER_TABLA,
            FECHA_HORA,
            IS_CONTIENE_BATERIA,
            IS_CONTIENE_GPS,
            IS_CONTIENE_TIEMPO_VIDA,
            IS_ERROR_OBTENER_BATERIA,
            IS_ERROR_OBTENER_GPS,
            IS_ERROR_OBTENER_TIEMPO_VIDA,
            LATITUD,
            LONGITUD,
            PORCENTAJE_BATERIA,
            TIEMPO_VIDA,
            FECHA_REGISTRO
        FROM USR_LAB.VW_ESTATUS_ZP_DJANGO
        WHERE AMID = :amid
          AND FECHA_HORA >= TO_DATE(:fecha_inicio, 'YYYY-MM-DD HH24:MI:SS')
          AND FECHA_HORA < TO_DATE(:fecha_fin, 'YYYY-MM-DD HH24:MI:SS')
        ORDER BY FECHA_HORA
    """

    registros = []

    with obtener_conexion_oracle() as conexion:
        with conexion.cursor() as cursor:
            cursor.execute(
                query,
                {
                    "amid": int(amid),
                    "fecha_inicio": fecha_inicio_texto,
                    "fecha_fin": fecha_fin_texto,
                },
            )

            columnas = [col[0].lower() for col in cursor.description]

            for fila in cursor.fetchall():
                datos = dict(zip(columnas, fila))
                registros.append(datos)

    return registros


def convertir_numero_excel(valor):
    """
    Convierte valores numéricos Oracle/Python a float.
    Sirve para cálculos del resumen.
    """

    if valor is None or valor == "":
        return None

    try:
        return float(valor)
    except (ValueError, TypeError):
        return None


def es_gps_cero(registro):
    """
    Detecta registros GPS 0,0.
    """

    latitud = convertir_numero_excel(registro.get("latitud"))
    longitud = convertir_numero_excel(registro.get("longitud"))

    return latitud == 0 and longitud == 0


def construir_resumen_exportacion(
    amid,
    registros_dict,
    fecha_inicio,
    fecha_fin,
    hora_inicio,
    hora_fin,
):
    """
    Construye indicadores generales para la hoja Resumen.
    """

    fechas = [
        preparar_valor_excel(registro.get("fecha_hora"))
        for registro in registros_dict
        if registro.get("fecha_hora")
    ]

    baterias = [
        convertir_numero_excel(registro.get("porcentaje_bateria"))
        for registro in registros_dict
        if convertir_numero_excel(registro.get("porcentaje_bateria")) is not None
    ]

    registros_con_gps = [
        registro
        for registro in registros_dict
        if registro.get("latitud") is not None
        and registro.get("longitud") is not None
    ]

    registros_gps_cero = [
        registro
        for registro in registros_con_gps
        if es_gps_cero(registro)
    ]

    ultimo_registro = registros_dict[-1] if registros_dict else None

    resumen = {
        "AMID": amid,
        "Periodo exportado": (
            f"{fecha_inicio.strftime('%d-%m-%Y %H:%M')} "
            f"a {fecha_fin.strftime('%d-%m-%Y %H:%M')}"
        ),
        "Horario tabla batería": f"{hora_inicio} a {hora_fin}",
        "Fecha exportación": obtener_ahora_referencia(),
        "Total registros completos": len(registros_dict),
        "Total registros con batería": len(baterias),
        "Total registros con GPS": len(registros_con_gps),
        "Total GPS 0,0": len(registros_gps_cero),
        "Primera fecha/hora registrada": min(fechas) if fechas else None,
        "Última fecha/hora registrada": max(fechas) if fechas else None,
        "Batería mínima": min(baterias) if baterias else None,
        "Batería máxima": max(baterias) if baterias else None,
        "Última batería registrada": preparar_valor_excel(
            ultimo_registro.get("porcentaje_bateria")
        ) if ultimo_registro else None,
        "Última latitud": preparar_valor_excel(
            ultimo_registro.get("latitud")
        ) if ultimo_registro else None,
        "Última longitud": preparar_valor_excel(
            ultimo_registro.get("longitud")
        ) if ultimo_registro else None,
    }

    return resumen


def obtener_diccionario_variables_exportacion():
    """
    Diccionario de variables de la hoja 'Registros completos'.
    Se muestra en la hoja Resumen para que el Excel sea autoexplicativo.
    """

    return [
        ("ID", "Identificador interno del registro en Oracle."),
        ("AMID", "Identificador del validador."),
        ("Fecha hora", "Fecha y hora del registro reportado por el validador."),
        ("Fecha descarga", "Última fecha de descarga registrada para el validador."),
        ("Fecha estado", "Última fecha de estado registrada para el validador."),
        ("BUSID", "Identificador del bus o equipo asociado, cuando existe."),
        ("OP", "Operador asociado al validador."),
        ("Versión", "Versión registrada del equipo o configuración."),
        ("Patente", "Patente asociada al bus, cuando existe."),
        ("TD01", "Contador o dato técnico TD01 informado por el equipo."),
        ("TD04", "Contador o dato técnico TD04 informado por el equipo."),
        ("Tabla", "Número de tabla de difusión consultada."),
        ("Ver tabla", "Versión de la tabla de difusión."),
        ("Latitud", "Latitud GPS reportada por el validador."),
        ("Longitud", "Longitud GPS reportada por el validador."),
        ("Porcentaje batería", "Porcentaje de batería reportado por el validador."),
        ("Tiempo vida", "Fecha/hora o valor asociado al tiempo de vida reportado por el equipo."),
        ("Fecha registro", "Fecha en que el dato fue registrado en la consulta o proceso."),
        ("Contiene batería", "Indica si el mensaje contenía información de batería."),
        ("Contiene GPS", "Indica si el mensaje contenía información GPS."),
        ("Contiene tiempo vida", "Indica si el mensaje contenía información de tiempo de vida."),
        ("Error obtener batería", "Indica si hubo error al obtener batería."),
        ("Error obtener GPS", "Indica si hubo error al obtener GPS."),
        ("Error obtener tiempo vida", "Indica si hubo error al obtener tiempo de vida."),
    ]


def crear_hoja_resumen(
    wb,
    amid,
    registros_dict,
    fecha_inicio,
    fecha_fin,
    hora_inicio,
    hora_fin,
):
    """
    Crea la hoja Resumen del Excel estándar.
    """

    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    resumen = construir_resumen_exportacion(
        amid=amid,
        registros_dict=registros_dict,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
    )

    ws_resumen.append(["Indicador", "Valor"])

    for indicador, valor in resumen.items():
        ws_resumen.append([
            indicador,
            preparar_valor_excel(valor),
        ])

    ws_resumen.append([])
    ws_resumen.append(["Notas", "Descripción"])
    ws_resumen.append([
        "Tabla batería",
        "Las celdas vacías significan que no hubo dato real para ese bloque horario. El valor 0 sí es un dato real reportado.",
    ])
    ws_resumen.append([
        "GPS 0,0",
        "Si Latitud = 0 y Longitud = 0, se considera GPS inválido reportado por el equipo, no ubicación real.",
    ])

    ws_resumen.append([])
    ws_resumen.append(["Diccionario de variables", ""])
    ws_resumen.append(["Variable", "Significado"])

    for variable, significado in obtener_diccionario_variables_exportacion():
        ws_resumen.append([variable, significado])

    aplicar_estilo_hoja(ws_resumen)

    ws_resumen.column_dimensions["A"].width = 34
    ws_resumen.column_dimensions["B"].width = 90

    for row in ws_resumen.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    return ws_resumen


def crear_excel_completo_amid(amid, dias=14, hora_inicio="00:00", hora_fin="23:30"):
    """
    Crea un Excel estándar para Baterías y GPS.

    Hoja 1: Resumen
    Hoja 2: Tabla batería
    Hoja 3: Registros completos
    """

    if not hora_inicio:
        hora_inicio = "00:00"

    if not hora_fin:
        hora_fin = "23:30"

    if hora_inicio > hora_fin:
        hora_inicio = "00:00"
        hora_fin = "23:30"

    dias = 14

    fecha_inicio, fecha_fin = obtener_rango_fechas_panel(dias)

    registros_dict = obtener_registros_completos_oracle(
        amid=amid,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    if not registros_dict:
        return None

    bloques_bateria = obtener_bloques_bateria_oracle(
        amid=amid,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )

    columnas_horas, tabla_bateria = construir_tabla_bateria(
        bloques=bloques_bateria,
        cantidad_dias=dias,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
    )

    wb = Workbook()

    # =========================
    # Hoja 1: resumen
    # =========================
    crear_hoja_resumen(
        wb=wb,
        amid=amid,
        registros_dict=registros_dict,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        hora_inicio=hora_inicio,
        hora_fin=hora_fin,
    )

    # =========================
    # Hoja 2: tabla batería
    # =========================
    ws_tabla = wb.create_sheet("Tabla bateria")

    ws_tabla.append(["AMID", amid])
    ws_tabla.append(["Periodo", "Últimos 14 días"])
    ws_tabla.append(["Horario", f"{hora_inicio} a {hora_fin}"])
    ws_tabla.append([])

    encabezados_tabla = ["Fecha"] + columnas_horas
    ws_tabla.append(encabezados_tabla)

    for fila in tabla_bateria:
        valores = [fila["fecha"]]

        for celda in fila["valores"]:
            valores.append(celda["valor"] if celda["valor"] != "" else None)

        ws_tabla.append(valores)

    aplicar_estilo_tabla_bateria(ws_tabla)

    # =========================
    # Hoja 3: registros completos
    # =========================
    ws_datos = wb.create_sheet("Registros completos")

    encabezados_datos = [
        "ID",
        "AMID",
        "Fecha hora",
        "Fecha descarga",
        "Fecha estado",
        "BUSID",
        "OP",
        "Versión",
        "Patente",
        "TD01",
        "TD04",
        "Tabla",
        "Ver tabla",
        "Latitud",
        "Longitud",
        "Porcentaje batería",
        "Tiempo vida",
        "Fecha registro",
        "Contiene batería",
        "Contiene GPS",
        "Contiene tiempo vida",
        "Error obtener batería",
        "Error obtener GPS",
        "Error obtener tiempo vida",
    ]

    ws_datos.append(encabezados_datos)

    for registro in registros_dict:
        ws_datos.append([
            preparar_valor_excel(registro.get("id")),
            preparar_valor_excel(registro.get("amid")),
            preparar_valor_excel(registro.get("fecha_hora")),
            preparar_valor_excel(registro.get("fec_descarga")),
            preparar_valor_excel(registro.get("fec_estado")),
            preparar_valor_excel(registro.get("busid")),
            preparar_valor_excel(registro.get("op")),
            preparar_valor_excel(registro.get("version")),
            preparar_valor_excel(registro.get("patente")),
            preparar_valor_excel(registro.get("td01")),
            preparar_valor_excel(registro.get("td04")),
            preparar_valor_excel(registro.get("tabla")),
            preparar_valor_excel(registro.get("ver_tabla")),
            preparar_valor_excel(registro.get("latitud")),
            preparar_valor_excel(registro.get("longitud")),
            preparar_valor_excel(registro.get("porcentaje_bateria")),
            preparar_valor_excel(registro.get("tiempo_vida")),
            preparar_valor_excel(registro.get("fecha_registro")),
            preparar_valor_excel(registro.get("is_contiene_bateria")),
            preparar_valor_excel(registro.get("is_contiene_gps")),
            preparar_valor_excel(registro.get("is_contiene_tiempo_vida")),
            preparar_valor_excel(registro.get("is_error_obtener_bateria")),
            preparar_valor_excel(registro.get("is_error_obtener_gps")),
            preparar_valor_excel(registro.get("is_error_obtener_tiempo_vida")),
        ])

    aplicar_estilo_hoja(ws_datos)

    return wb


@login_required
def panel_perfil(request):
    usuario_actual = request.user

    es_admin = usuario_es_admin(usuario_actual)

    roles_usuario_actual = usuario_actual.groups.values_list("name", flat=True)
    roles_usuario_actual = ", ".join(roles_usuario_actual)

    context = {
        "active_page": "perfil",
        "es_admin": es_admin,
        "usuario_actual": usuario_actual,
        "roles_usuario_actual": roles_usuario_actual or "Sin rol asignado",
        "abrir_editor_reglas": request.GET.get("editor_reglas") == "1",
    }

    if request.method == "POST" and es_admin:
        accion = request.POST.get("action")

        if accion in {"guardar_reglas_alertas", "guardar_y_recalcular_alertas"}:
            try:
                resultado = actualizar_reglas_alertas(request.POST)
                cantidad = resultado["cantidad"]
                modo_recalculo = resultado["modo_recalculo"]

                if cantidad == 0:
                    messages.info(request, "No se detectaron cambios en las reglas.")
                elif accion == "guardar_y_recalcular_alertas":
                    hilo = iniciar_recalculo_en_segundo_plano(
                        modo_recalculo=modo_recalculo
                    )
                    if hilo is None:
                        messages.warning(
                            request,
                            f"Se guardaron {cantidad} regla(s), pero ya hay un "
                            "recálculo manual en curso. No se inició otro proceso.",
                        )
                    else:
                        descripcion_modo = (
                            "completo, porque cambió una regla de detección"
                            if modo_recalculo == "completo"
                            else "rápido, porque solo cambiaron reglas de clasificación"
                        )
                        messages.success(
                            request,
                            f"Se guardaron {cantidad} regla(s). Se inició el recálculo "
                            f"{descripcion_modo} en segundo plano.",
                        )
                else:
                    messages.success(
                        request,
                        f"Se guardaron {cantidad} regla(s) y se validaron en Oracle.",
                    )
            except ValueError as error:
                messages.error(request, str(error))
            except Exception as error:
                messages.error(request, f"No fue posible actualizar las reglas: {error}")

            url_perfil = reverse("dashboard:panel_perfil")
            return redirect(f"{url_perfil}?editor_reglas=1")

    if es_admin:
        usuarios = User.objects.all().order_by("username")

        total_usuarios = usuarios.count()
        usuarios_activos = usuarios.filter(is_active=True).count()
        usuarios_inactivos = usuarios.filter(is_active=False).count()
        total_admins = usuarios.filter(is_superuser=True).count()
        ultimos_logs = LogImportacion.objects.all().order_by("-fecha_inicio")[:8]
        resultado_comando_admin = request.session.pop("resultado_comando_admin", None)
        grupos = Group.objects.all().order_by("name")

        resumen_roles = []

        for grupo in grupos:
            resumen_roles.append({
                "nombre": grupo.name,
                "cantidad": grupo.user_set.count(),
            })

        lista_usuarios = []

        for usuario in usuarios:
            grupos_usuario = usuario.groups.all()
            roles = ", ".join([grupo.name for grupo in grupos_usuario])

            if usuario.is_superuser:
                roles = "Admin" if not roles else f"Admin, {roles}"

            lista_usuarios.append({
                "username": usuario.username,
                "email": usuario.email,
                "nombre": usuario.get_full_name() or "-",
                "roles": roles or "Sin rol",
                "activo": usuario.is_active,
                "ultimo_acceso": usuario.last_login,
                "fecha_creacion": usuario.date_joined,
                "es_admin": usuario.is_superuser,
            })

        context.update({
            "total_usuarios": total_usuarios,
            "usuarios_activos": usuarios_activos,
            "usuarios_inactivos": usuarios_inactivos,
            "total_admins": total_admins,
            "resumen_roles": resumen_roles,
            "lista_usuarios": lista_usuarios,
            "ultimos_logs": ultimos_logs,
            "resultado_comando_admin": resultado_comando_admin,
            "puede_editar_reglas": usuario_puede_editar_reglas(usuario_actual),
        })

    return render(request, "dashboard/panel_perfil.html", context)


@login_required
def editor_reglas_alertas(request):
    """Renderiza el editor bajo demanda para no consultar Oracle al abrir el perfil."""
    if request.method != "GET":
        return HttpResponse("Método no permitido.", status=405)

    if not usuario_puede_editar_reglas(request.user):
        return HttpResponse("No tienes permisos para editar estas reglas.", status=403)

    contexto = {
        "editor_reglas": {"BATERIA": [], "GPS": []},
        "recalculo_en_curso": recalculo_en_curso(),
        "log_recalculo": leer_log_recalculo(),
    }
    estado = 200
    try:
        editor = obtener_editor_reglas_alertas()
        contexto["editor_reglas"] = editor
        contexto["total_reglas_editor"] = sum(
            len(seccion["reglas"])
            for secciones in editor.values()
            for seccion in secciones
        )
    except Exception as error:
        contexto["reglas_alertas_error"] = str(error)
        estado = 503

    return render(
        request,
        "dashboard/partials/editor_reglas_alertas.html",
        contexto,
        status=estado,
    )


@login_required
def exportar_baterias_excel(request):
    """
    Exporta el Excel estándar del AMID desde el botón Baterías.

    Por ahora:
    - siempre exporta 14 días completos
    - horario completo 00:00 a 23:30
    """

    amid = request.GET.get("amid", "").strip()

    if not amid:
        return HttpResponse("Debe indicar un AMID para exportar.", status=400)

    dias = 14
    hora_inicio = "00:00"
    hora_fin = "23:30"

    try:
        wb = crear_excel_completo_amid(
            amid=amid,
            dias=dias,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin,
        )
    except ValueError:
        return HttpResponse("El AMID ingresado no es válido.", status=400)
    except Exception as error:
        return HttpResponse(
            f"Error consultando datos en Oracle: {error}",
            status=500,
        )

    if wb is None:
        return HttpResponse(
            f"No existen registros para el AMID {amid} en los últimos 14 días.",
            status=404,
        )

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    fecha_exportacion = obtener_ahora_referencia().strftime("%Y%m%d_%H%M%S")
    filename = f"datos_amid_{amid}_14_dias_{fecha_exportacion}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def aplicar_estilo_hoja(ws):
    """
    Aplica formato básico a una hoja Excel.
    """

    fill_header = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid",
    )

    font_header = Font(
        color="FFFFFF",
        bold=True,
    )

    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = min(max_length + 2, 35)
        ws.column_dimensions[column_letter].width = adjusted_width


def preparar_valor_excel(valor):
    """
    Prepara valores para Excel.
    Si es datetime con zona horaria, lo convierte a hora local
    y elimina tzinfo, porque Excel no soporta timezone.
    """

    if valor is None:
        return None

    if isinstance(valor, datetime):
        if timezone.is_aware(valor):
            valor = timezone.localtime(valor)

        return valor.replace(tzinfo=None)

    return valor


def aplicar_estilo_tabla_bateria(ws):
    """
    Aplica estilo a la hoja con la matriz fecha x hora.
    """

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    font_titulo = Font(color="FFFFFF", bold=True)
    font_header = Font(bold=True, color="1F1F1F")

    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    # Estilo de las primeras filas informativas
    for row in range(1, 4):
        ws[f"A{row}"].fill = fill_titulo
        ws[f"A{row}"].font = font_titulo
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

    # Encabezados están en la fila 5
    fila_header = 5

    for cell in ws[fila_header]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Bordes + alineación para toda la tabla
    for row in ws.iter_rows(min_row=fila_header, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Colores de batería en celdas numéricas
    for row in ws.iter_rows(
        min_row=fila_header + 1,
        max_row=ws.max_row,
        min_col=2,
        max_col=ws.max_column,
    ):
        for cell in row:
            valor = cell.value

            if valor is None:
                cell.fill = PatternFill("solid", fgColor="F1F3F5")
                continue

            try:
                bateria = float(valor)
            except (ValueError, TypeError):
                continue

            if bateria >= 80:
                cell.fill = PatternFill("solid", fgColor="D1E7DD")
            elif bateria >= 50:
                cell.fill = PatternFill("solid", fgColor="FFF3CD")
            elif bateria >= 20:
                cell.fill = PatternFill("solid", fgColor="FFE5D0")
            else:
                cell.fill = PatternFill("solid", fgColor="F8D7DA")

    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"A5:{get_column_letter(ws.max_column)}{ws.max_row}"

    ws.column_dimensions["A"].width = 14

    for col in range(2, ws.max_column + 1):
        letra = get_column_letter(col)
        ws.column_dimensions[letra].width = 8
